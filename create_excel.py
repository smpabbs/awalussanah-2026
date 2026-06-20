import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

with open('X:/matrikulasi 2627/data/siswa.json') as f:
    data = json.load(f)

siswa = sorted(data['siswa'], key=lambda x: (x['kelas'], x['n']))
by_kelas = defaultdict(list)
for s in siswa:
    by_kelas[s['kelas']].append(s)

walas = {
    "7 Putra A": "Daffa Danendra Rizqi Nugraha, S.Pd.",
    "7 Putra B": "Hang Sakti Abdullah, S.Pd.",
    "7 Putra C": "Ifan Destya Adi Tama, S.Pd, Gr",
    "7 Putri A": "Aulia Qisthi Rosyada, S.Pd., Gr",
    "7 Putri B": "Charlieta Nova Putri Fedito, S.Kom."
}

wb = Workbook()

# Colors
gold_fill = PatternFill(start_color="C8A45C", end_color="C8A45C", fill_type="solid")
dark_fill = PatternFill(start_color="0A0E17", end_color="0A0E17", fill_type="solid")
teal_fill = PatternFill(start_color="2DD4BF", end_color="2DD4BF", fill_type="solid")
header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="C8A45C")
normal_font = Font(name="Calibri", size=11, color="333333")
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# ===== Sheet 1: Summary =====
ws_summary = wb.active
ws_summary.title = "Ringkasan"
ws_summary.sheet_properties.tabColor = "C8A45C"

# Title
ws_summary.merge_cells('A1:E1')
ws_summary['A1'] = "PEMBAGIAN KELAS SMP ABBS SURAKARTA"
ws_summary['A1'].font = Font(name="Calibri", size=18, bold=True, color="C8A45C")
ws_summary['A1'].alignment = Alignment(horizontal='center')

ws_summary.merge_cells('A2:E2')
ws_summary['A2'] = "Matrikulasi TA 2026/2027"
ws_summary['A2'].font = Font(name="Calibri", size=12, color="666666")
ws_summary['A2'].alignment = Alignment(horizontal='center')

# Summary table
row = 4
headers = ["Kelas", "Wali Kelas", "Laki-laki", "Perempuan", "Total"]
for col, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = gold_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row = 5
total_l = 0
total_p = 0
for kelas in sorted(by_kelas.keys()):
    siswa_kelas = by_kelas[kelas]
    l_count = sum(1 for s in siswa_kelas if s['g'] == 'L')
    p_count = sum(1 for s in siswa_kelas if s['g'] == 'P')
    total_l += l_count
    total_p += p_count
    
    ws_summary.cell(row=row, column=1, value=kelas).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=walas.get(kelas, '-'))
    ws_summary.cell(row=row, column=3, value=l_count).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=4, value=p_count).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=5, value=len(siswa_kelas)).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=5).font = Font(bold=True)
    
    for col in range(1, 6):
        ws_summary.cell(row=row, column=col).border = thin_border
    row += 1

# Total row
ws_summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, color="C8A45C")
ws_summary.cell(row=row, column=3, value=total_l).alignment = Alignment(horizontal='center')
ws_summary.cell(row=row, column=3).font = Font(bold=True)
ws_summary.cell(row=row, column=4, value=total_p).alignment = Alignment(horizontal='center')
ws_summary.cell(row=row, column=4).font = Font(bold=True)
ws_summary.cell(row=row, column=5, value=len(siswa)).alignment = Alignment(horizontal='center')
ws_summary.cell(row=row, column=5).font = Font(bold=True, color="C8A45C", size=12)
for col in range(1, 6):
    ws_summary.cell(row=row, column=col).border = thin_border
    ws_summary.cell(row=row, column=col).fill = PatternFill(start_color="FFF8E7", end_color="FFF8E7", fill_type="solid")

# Column widths
ws_summary.column_dimensions['A'].width = 15
ws_summary.column_dimensions['B'].width = 40
ws_summary.column_dimensions['C'].width = 12
ws_summary.column_dimensions['D'].width = 12
ws_summary.column_dimensions['E'].width = 10

# ===== Sheet per Kelas =====
for kelas in sorted(by_kelas.keys()):
    safe_name = kelas.replace(' ', '_')
    ws = wb.create_sheet(title=safe_name)
    ws.sheet_properties.tabColor = "2DD4BF" if "Putri" in kelas else "C8A45C"
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "KELAS " + kelas.upper()
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = "Wali Kelas: " + walas.get(kelas, '-')
    ws['A2'].font = Font(name="Calibri", size=11, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    row = 4
    headers = ["No", "Nama Siswa", "Jenis Kelamin", "Asal Sekolah", "No. WA Orang Tua", "Keterangan"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = gold_fill if "Putra" in kelas else teal_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    row = 5
    for i, s in enumerate(by_kelas[kelas], 1):
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=s['n'])
        gender = "Laki-laki" if s['g'] == 'L' else "Perempuan"
        ws.cell(row=row, column=3, value=gender).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=s['s'])
        ws.cell(row=row, column=5, value="")  # No WA placeholder
        ws.cell(row=row, column=6, value="")  # Keterangan placeholder
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = normal_font
        
        # Alternating row color
        if i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        row += 1
    
    # Summary row
    ws.cell(row=row, column=1, value="").border = thin_border
    ws.cell(row=row, column=2, value="Total: " + str(len(by_kelas[kelas])) + " siswa").font = Font(bold=True, color="C8A45C")
    ws.cell(row=row, column=2).border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

# Save
wb.save('X:/matrikulasi 2627/Pembagian_Kelas_2026.xlsx')
print("Excel saved: Pembagian_Kelas_2026.xlsx")
print("Total siswa:", len(siswa))
print("Kelas:", sorted(by_kelas.keys()))
